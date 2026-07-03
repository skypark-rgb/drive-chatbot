from io import BytesIO

from openpyxl import load_workbook

from models.document import Document


class ExcelParser:
    def __init__(self, drive_client):
        self.drive_client = drive_client

    def parse(self, file_metadata):
        excel_bytes = self.drive_client.download_file(
            file_metadata.id
        )

        workbook = load_workbook(
            filename=BytesIO(excel_bytes),
            data_only=True,
        )

        text_lines = []

        for sheet in workbook.worksheets:
            text_lines.append(f"Sheet: {sheet.title}")

            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cleaned_cells = [
                    str(cell).strip()
                    for cell in row
                    if cell is not None and str(cell).strip()
                ]

                if not cleaned_cells:
                    continue

                text_lines.append(
                    f"Row {row_index}: " + " | ".join(cleaned_cells)
                )

            text_lines.append("")

        return Document(
            file_id=file_metadata.id,
            name=file_metadata.name,
            mime_type=file_metadata.mime_type,
            text="\n".join(text_lines),
        )